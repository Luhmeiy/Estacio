import openpyxl
import Banco_de_dados
import tkinter
import tkinter.messagebox


class funcoes(Banco_de_dados.BD):
    def verificar_login(self, event):
        BD = self.abrir_BD()
        ws = BD['logins']

        for lin in range(2, 1_000_000, 1):
            if ws.cell(row=lin, column=1).value is None:
                tkinter.messagebox.showinfo('Login', 'Matricula não existe!')
                break
            if ws.cell(row=lin, column=1).value == self.usuario.entry.get():
                if ws.cell(row=lin, column=2).value == self.senha.entry.get():
                    tkinter.messagebox.showinfo('Login', 'Login correto')
                else:
                    tkinter.messagebox.showinfo('Login', 'Senha errada!')
                break
        self.fechar_BD(BD)
